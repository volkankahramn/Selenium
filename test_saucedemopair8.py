
import pytest
import time
import json
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from pathlib import Path
from constants import *
import openpyxl


class TestSaucedemopair8():
  def setup_method(self, method):
    self.driver = webdriver.Chrome()
    self.vars = {}
  
  def teardown_method(self, method):
    self.driver.quit()
  
  def test_lockedoutuser(self): #locked_out_user ile giriş yapıldığında verilen uyarı mesajının doğrulanması.
    self.driver.get(BASE_DOMAIN_URL)
    self.driver.set_window_size(1552, 848)
    WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, USERNAME)))
    self.driver.find_element(By.CSS_SELECTOR, USERNAME).click()
    self.driver.find_element(By.CSS_SELECTOR, USERNAME).send_keys("locked_out_user")
    WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, PASSWORD)))
    self.driver.find_element(By.CSS_SELECTOR, PASSWORD).click()
    self.driver.find_element(By.CSS_SELECTOR, PASSWORD).send_keys("secret_sauce")
    WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, LOGINBTN)))
    self.driver.find_element(By.CSS_SELECTOR, LOGINBTN).click()
    WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, ERROR)))
    assert self.driver.find_element(By.CSS_SELECTOR, ERROR).text == "Epic sadface: Sorry, this user has been locked out."
    self.driver.close()
    
  def test_login(self): #normal giriş yapılıp sonrasında ürünlerin sayısının doğrulanması
        
        self.driver.get(BASE_DOMAIN_URL)
        self.driver.maximize_window()
        WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, USERNAME)))
        username = self.driver.find_element(By.CSS_SELECTOR, USERNAME)
        username.send_keys("standard_user")
        WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, PASSWORD)))
        password = self.driver.find_element(By.CSS_SELECTOR, PASSWORD)
        password.send_keys("secret_sauce")
        loginBtn = self.driver.find_element(By.CSS_SELECTOR, LOGINBTN)
        loginBtn.click()

  def test_inventor_item(self): #normal giriş yapılıp sonrasında ürünlerin sayısının doğrulanması
        
        self.test_login()

        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.XPATH, INVENTORY_ITEM)))
        image = self.driver.find_elements(By.XPATH, INVENTORY_ITEM)
        assert len(image)  == 6 #6 ürünün doğrulunu assert kodu.
        

        #ürünlerin isimlerinin excel dosyalarındaki isimlerle uyuşması
  def readProductNameFromExcel():
        excelFile = openpyxl.load_workbook("data/test_product_name.xlsx")
        selectedSheet = excelFile["Sheet1"]
        
        rows = selectedSheet.max_row
        data = []

        for i in range(2,rows+1):
            product_name = selectedSheet.cell(i,1).value 
            tupleExample = (product_name)
            data.append(tupleExample)
        return data

  @pytest.mark.parametrize("product_name",readProductNameFromExcel())
  def test_product_check(self,product_name):
    self.test_login()
    WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, INVENTORY_ITEMNAME)))
    assert self.driver.find_element(By.CSS_SELECTOR, INVENTORY_ITEMNAME).text == "Sauce Labs Backpack"
    self.driver.close()


    #bir excel dosyasında ismi geçen ürünlerin sepete eklenmesi fonksiyonu testi
  def readProductNameFromExcel():
        excelFile = openpyxl.load_workbook("data/test_product_name.xlsx")
        selectedSheet = excelFile["Sheet1"]
        
        rows = selectedSheet.max_row
        data = []

        for i in range(2,rows+1):
            product_add = selectedSheet.cell(i,1).value 
            tupleExample = (product_add)
            data.append(tupleExample)
        return data
  @pytest.mark.parametrize("product_add",readProductNameFromExcel())
  def test_productadd(self,product_add):
    self.test_login()
    WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, INVENTORY_ITEMNAME)))
    self.driver.find_element(By.CSS_SELECTOR, ADD_TO_CART).click()
    self.driver.find_element(By.CSS_SELECTOR, BASKET).click()
    WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, INVENTORY_ITEM1)))
    elements = self.driver.find_elements(By.CSS_SELECTOR, INVENTORY_ITEM1)
    assert len(elements) > 0
    self.driver.close()

  #sepete eklenen ürünlerin sepet sayfasında doğru bir şekilde görünmesi testi
  def test_baskets(self):
    self.test_login()
    WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, INVENTORY_ITEMNAME)))
    self.driver.find_element(By.CSS_SELECTOR, ADD_TO_CART).click()
    self.driver.find_element(By.CSS_SELECTOR, ADD_TO_CART1).click()
    self.driver.find_element(By.CSS_SELECTOR, ADD_TO_CART2).click()
    WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, BASKET)))
    self.driver.find_element(By.LINK_TEXT, "3").click()
    WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, ".cart_desc_label")))
    assert self.driver.find_element(By.CSS_SELECTOR, BASKET).text == "3"
    self.driver.close()

    #sepetten kaldırılan ürünün sepet ekranından kaldırılma testi
  def test_basketremove(self):
    self.test_login()
    self.driver.find_element(By.CSS_SELECTOR, ADD_TO_CART).click()
    self.driver.find_element(By.CSS_SELECTOR, ADD_TO_CART1 ).click()
    self.driver.find_element(By.CSS_SELECTOR, ADD_TO_CART2).click()
    WebDriverWait(self.driver, 3).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, BASKET)))
    self.driver.find_element(By.CSS_SELECTOR, BASKET).click()
    WebDriverWait(self.driver, 3).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, INVENTORY_ITEMNAME)))
    assert self.driver.find_element(By.CSS_SELECTOR, BASKET).text == "3"
    self.driver.find_element(By.CSS_SELECTOR, REMOVE_TO_CART).click()
    self.driver.find_element(By.CSS_SELECTOR, REMOVE_TO_CART1).click()
    self.driver.find_element(By.CSS_SELECTOR, REMOVE_TO_CART2).click()
    WebDriverWait(self.driver, 3).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, BASKET_LINK)))
    elements = self.driver.find_elements(By.CSS_SELECTOR, BASKET_LINK)
    assert len(elements) > 0
    self.driver.close()
      
  #Ürünleri Z'den A'ya sıralama
  def test_productz2a(self): #normal giriş yapılıp sonrasında ürünlerin sayısının doğrulanması
        
        self.test_login()

        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.XPATH, INVENTORY_ITEM )))
        image = self.driver.find_elements(By.XPATH, INVENTORY_ITEM )
        assert len(image)  == 6 #6 ürünün doğrulunu assert kodu.
        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.CLASS_NAME, PRODUCT_SORT)))
        product_arrangement = self.driver.find_element(By.CLASS_NAME, PRODUCT_SORT)
        product_arrangement.click()

        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.XPATH, VALUE)))
        product_z2a = self.driver.find_element(By.XPATH, VALUE)
        product_z2a.click()

        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.XPATH, INVENTORY_ITEM)))
        product = self.driver.find_elements(By.XPATH, INVENTORY_ITEM)
        product_z2a = self.driver.find_elements(By.XPATH, INVENTORY_ITEM) 
        productName = len(product)
    
        List = []
        for i in range (0, productName): 
          product_name = product_z2a[i]  
          productText = product_name.text 
          List.append(productText) 
          List = sorted(List, reverse = True)       
        
        return List

#Ürünleri düşük fiyattan yüksek fiyata listeleme

  def test_productl2h(self): #normal giriş yapılıp sonrasında ürünlerin sayısının doğrulanması
        
        self.test_login()

        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.XPATH, INVENTORY_ITEM )))
        image = self.driver.find_elements(By.XPATH, INVENTORY_ITEM )
        assert len(image)  == 6 #6 ürünün doğrulunu assert kodu.

        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.CLASS_NAME, PRODUCT_SORT)))
        product_arrangement = self.driver.find_element(By.CLASS_NAME, PRODUCT_SORT)
        product_arrangement.click()

        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.XPATH, VALUELOW)))
        product_l2h = self.driver.find_element(By.XPATH, VALUELOW)
        product_l2h.click()

        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.XPATH, INVENTORY_ITEM)))
        price = self.driver.find_elements(By.XPATH, INVENTORY_ITEM)
        product_cost = self.driver.find_elements(By.XPATH, PRICE) 
        productL2hlen = len(price)
    
        List = []
        for i in range (0, productL2hlen):
          product_price = product_cost[i] 
          product_itemNumber = float(product_price.text.replace("$",""))
          List.append(product_itemNumber)
          List=sorted(List)
        
        return List


